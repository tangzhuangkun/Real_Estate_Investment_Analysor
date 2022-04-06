import xlrd
import openpyxl

class Cal:
	def __init__(self):
		'''
		asking for method to calculate payback
		and requiring correct choice
		
		1st: cal by file
			1st: cal whole file to get the payback period(yearly and monthly) data;
			2nd: cal only one row of the file to get this row's payback period(yearly and monthly) data
		2nd: cal by single sample
		'''
		
		print("读取文件还是计算单个案例回报率：a:文件 b:单个")
		choose = input()
		while choose != "a" and choose != "b":
			print("只能选a或b")
			print("读取文件还是计算单个回报率：a:文件 b:单个")
			choose = input()
			
		else:
			if	choose == "a":
				print("计算整个文件还是计算单行：1:文件 2:单行")
				fileOrPiece = input()
				while fileOrPiece != "1" and fileOrPiece != "2":
					print("只能选1或2")
					print("计算整个文件还是计算单行：1:文件 2:单行")
					fileOrPiece = input()
				else:
					if fileOrPiece == "1":
						self.calWholePaybackFile()
					else:
						self.calOneFileRowData()
				
			elif choose == "b":
				self.calPaybackMonthes()
				
	
	def calOneFileRowData(self):
		'''
		read the excel data
		cal the investment, rent, rate
		to get the totalYear, totalMonth for particular row's result
		'''
		
		# read
		workbook = openpyxl.load_workbook('WaitingList.xlsx')
		#infoSheet = workbook.worksheets[0]   # this command same as the below
		infoSheet = workbook["Info"] 
		print("哪一行： ")
		parRowStr = input()
		parRow = int(parRowStr)
		
		print("地址："+infoSheet.cell(parRow,1).value)
		investmentStr = infoSheet.cell(parRow,2).value
		rentStr = infoSheet.cell(parRow,3).value 
		rateStr = infoSheet.cell(parRow,4).value 
		rent = float(rentStr)
		rate = float(rateStr)
		investment = float(investmentStr)
		totalYear,totalMonth = self.calRentChange(rent, rate, investment)

		
	
	def calWholePaybackFile(self):
		'''
		read the excel data and modify
		cal the investment, rent, rate
		to get the totalYear, totalMonth for each row's data
		then write the result to respective cell
		'''
		
		# read
		workbook = openpyxl.load_workbook('WaitingList.xlsx')
		#infoSheet = workbook.worksheets[0]   # this command same as the below
		infoSheet = workbook["Info"] 

			 
		nrows = infoSheet.max_row
		for i in range(2,nrows+1):
			print("地址："+infoSheet.cell(i,1).value)
			investmentStr = infoSheet.cell(i,2).value
			rentStr = infoSheet.cell(i,3).value 
			rateStr = infoSheet.cell(i,4).value 
			rent = float(rentStr)
			rate = float(rateStr)
			investment = float(investmentStr)
			totalYear,totalMonth = self.calRentChange(rent, rate, investment)
			print()
			print()
			
			# write data to respective cell	
			infoSheet.cell(i,10,totalYear)
			infoSheet.cell(i,11,totalMonth)
		
		# save the modified form 	
		workbook.save('WaitingList.xlsx')	
			
			
	def askForInput(self):
		
		'''
		asking for input
		'''		
		print("月租金(万)： ")
		rentStr = input()
		rent = float(rentStr)
		
		print("年增长率(小数)： ")
		rateStr = input()
		rate = float(rateStr)
		
		print("总投资(万)： ")
		investmentStr = input()
		investment = float(investmentStr)
		
		return rent,rate,investment
	
	
	
	def calRentChange(self,rent,rate,investment):
		'''
		cal each month rent
		cal each year rent income
		cal total rent income 
		cal payback period by years and monthes
		'''
		
		totalRent = 0
		totalMonth = 0
		totalYear = -1 
			
		for i in range(1000):
			j = i//12
			
			monthRent = rent*((1+rate)**j)
			yearRent = rent*((1+rate)**j)*12
			totalRent += monthRent
			totalMonth += 1
			
			if totalYear<j:
				totalYear = j
				print("第"+str(totalYear+1)+"年")
				print("月度租金收入: " + str(round(monthRent,4)))
				print("年度租金收入: " + str(round(yearRent,4)))
				print("租金总收入: " + str(round(totalRent-monthRent+yearRent,4)))
			
			if totalRent > investment:
					break
		print("共需" + str(totalYear+1)+ "年回本")			
		print("共需" + str(totalMonth)+ "个月回本")	
		return totalYear+1,totalMonth
		
			
	def calPaybackMonthes(self):
		'''
		print out each yearly and monthly change
		'''
				
		rent,rate,investment = self.askForInput()
		self.calRentChange(rent,rate,investment)
		
					
result = Cal()			

		