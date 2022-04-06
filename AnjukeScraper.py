import urllib
from urllib import request
from bs4 import BeautifulSoup            #Beautiful Soup是一个可以从HTML或XML文件中提取结构化数据的Python库
import ssl # 解决访问Https时不受信任SSL证书问题
import re #正则表达式
import openpyxl    #读写excel 
from urllib.parse import quote #把中文字符按照浏览器机制进行编码
import string
import GetProxyIP  #引入代理IP
import random
import Disguise

'''
version: 201900607.1832



功能表：
done:
2018.12.13: 每半小时抓取一次;  ------done
2018.12.13：URL里，中文转化为正确的链接输入;	------done
2018.12.16：通过关键词或目标地点定时运行; ------done
2019.01.27：使用代理IP和轮换  ----done
2019.01.27：user_agent 伪装 ---done


waited to be added
2019.01.27：设置一个随机时间间隔
2019.01.27：使用了代理IP和头文件伪装，仍然被封
2019.01.06: 如果‘预期收益’为空，则挖掘详细信息中的关键词，月租，年租，租金，递增;
2019.01.06：挖取关键词的条件，总价低于等于120万，面积小于等于80平米；
2019.01.13：设置访问时间间隔
2019.01.13：图形化界面
2019.02.27:抓取数据缓存之后再写入文件
2019.04.23:抓取，缓存后，存储，独立分析


''' 

class AnjukeSpider:
	
	
	global rawProxyIPList 
	rawProxyIPList = Disguise.Disguise().getRawIPList()
	print("rawProxyIPList--------")
	print()
	print(rawProxyIPList)
	print()

	
	def __init__(self):
		#解决访问Https时不受信任SSL证书问题
		ssl._create_default_https_context = ssl._create_unverified_context
		
	def loadLocalData(self):
		
		#读取本地文件，加载所有已经爬取过的门面信息的链接
		localLinksSet = set()
		wb = openpyxl.load_workbook('安居客爬取数据.xlsx')
		dataSheet = wb["重庆"]
		nrows = dataSheet.max_row
		for i in range(2,nrows+1): 
			localLinksSet.add(dataSheet.cell(i,13).value)
		#print(localLinksSet)
		return localLinksSet
		
	def checkLinkExistOrNot(self, link, localLinksSet):
		#检查当前link是否已抓取保存在本地，
		#存在，返回True；
		#不存在，返回False。
		if link in localLinksSet:
			return True
		else:
			localLinksSet.add(link)
			return False
		

	def goToMainPage(self,pageNum,location="",keyword=''):
		'''
		go to individual main page and get all main page info 
		'''
		#构造头文件，模拟浏览器访问
		rawURL = "https://cq.sp.anjuke.com/shou/"+location+"-p"+str(pageNum)+"/?kw="+keyword
		#把中文字符按照浏览器机制进行编码
		url  = quote(rawURL,safe=string.printable)
		print(url)
		
		#不使用代理IP
		#page = request.Request(url,headers=header)
		#page_info = request.urlopen(page).read().decode('utf-8')#打开Url,获取HttpResponse返回对象并读取其ResposneBody
		
		#使用封装方法得到一个假的头文件
		fakeHeader = Disguise.Disguise().getAFakeHeader()
		#print("main page using header: ", fakeHeader )
		
		#随机取得一个假的头文件
		#fakeHeader=random.choice(fakeHeadersList)
		
		#使用封装方法调用代理IP
		page_info = Disguise.Disguise().requestByProxyIP(url,fakeHeader,rawProxyIPList)		

		
		# 将获取到的内容转换成BeautifulSoup格式，并将html.parser作为解析器
		soup = BeautifulSoup(page_info, 'html.parser')

		# 以格式化的形式打印html
		##print(soup.prettify())

		#找到主页面中关于目标物业的关键信息，并返回beautifulsoup的一个列表
		list_items = soup.find_all('div', 'list-item')

		return list_items
		
		
		
	def getWholePageNewInfo(self,list_items,localLinksSet):
		#变量beautifulsoup返回的一个列表，得到每一个物业的链接
		#返回整页的所有有效且新链接的信息列表
		
		#大list套小list
		#大list存储整页所有有效连接中的详细信息
		#小list存储详细页中的7个关键信息
		wholePageLinklinkDetailInfoRows = []
		
		for item in list_items:
			##print(item)
			#get link
			link = item.get('link')
			print(link)
			
			#link是否存在在本地了
			#存在，不再抓取；
			#不存在，抓取
			exist = self.checkLinkExistOrNot(link,localLinksSet)
			if not exist:
				linkDetailInfoRow = self.analyzeLinkDetailInfo(link)
				wholePageLinklinkDetailInfoRows.append(linkDetailInfoRow)
	
		#print(wholePageLinklinkDetailInfoRows)
		return wholePageLinklinkDetailInfoRows
		
		
	def writeWholePageInfoAndSave(self,rows):
		#传入整页的所有有效且新链接的信息列表
		#逐行写本地存储文档中
		
		#打开本地存储文档
		wb = openpyxl.load_workbook('安居客爬取数据.xlsx')
		dataSheet = wb["重庆"]
		maxRow = dataSheet.max_row
		
		for i in range(len(rows)):
			#该行关键信息存储顺序为，链接，大概地理位置，总价，单价，预期租金收益，上传日期，备注信息，面积
			#i从0开始计数，但写入sheet时，需要跳一行，因此写成 maxRow+1+i
			
			dataSheet.cell(maxRow+1+i,13,rows[i][0])
			dataSheet.cell(maxRow+1+i,1,rows[i][1])
			dataSheet.cell(maxRow+1+i,2,rows[i][2])
			dataSheet.cell(maxRow+1+i,5,rows[i][3])
			dataSheet.cell(maxRow+1+i,3,rows[i][4])
			dataSheet.cell(maxRow+1+i,7,rows[i][5])
			dataSheet.cell(maxRow+1+i,12,rows[i][6])
			dataSheet.cell(maxRow+1+i,6,rows[i][7])
			
		wb.save('安居客爬取数据.xlsx')
		
			
	def analyzeLinkDetailInfo(self,link):
		
		
		#解析链接，去到每一个相关子界面parse link and go to each child page
		#不使用代理IP
		#page_child = request.Request(link,headers=header)
		#page_child_info = request.urlopen(page_child).read().decode('utf-8')#打开Url,获取HttpResponse返回对象并读取其ResposneBody
		
		#构建一个行，存储每一个店面详细页的关键信息，以便写入Excel表格
		#该行关键信息存储顺序为，链接，大概地理位置，总价，单价，预期租金收益，上传日期，备注信息，面积
		linkDetailInfoRow=['','','','','','','','']
		
		#使用封装方法得到一个假的头文件
		fakeHeader = Disguise.Disguise().getAFakeHeader()
		#print("child page using header: ", fakeHeader )
		
		#随机取得一个假的头文件
		#fakeHeader=random.choice(fakeHeadersList)
		
		#使用封装方法调用代理IP
		page_child_info = Disguise.Disguise().requestByProxyIP(link,fakeHeader,rawProxyIPList)
		 
		# 并将html.parser作为解析器
		soup_childPage = BeautifulSoup(page_child_info, 'html.parser')
		
		#解析子界面中的内容，得到相关物业的关键信息
		childPage_detail_info = soup_childPage.find('div','itemCon clearfix nocopy fy_info')
		
		
		#链接写入文件中
		#dataSheet.cell(maxRow+1,13,link)
		
		#链接写入关键信息行
		linkDetailInfoRow[0]=link
		
		#得到大概的地址信息，写入文件中
		rough_address = childPage_detail_info.find('span','desc addresscommu').get('title')
		#dataSheet.cell(maxRow+1,1,str(rough_address))
		
		#大概地理位置信息写入关键信息行
		linkDetailInfoRow[1]=rough_address
		
		all_nodes = childPage_detail_info.find_all('span','fst')
		total_price = ""
		unit_price  = ""
		rent        = ""
		
		for node in all_nodes:
			#得到总价，转化为float，写入文件中
			if node.get_text()=='总价：':
				total_price = node.find_next_sibling().get_text()
				total_price = "".join(total_price.split())
				#dataSheet.cell(maxRow+1,2,total_price)
				#总价写入关键信息行
				linkDetailInfoRow[2]=total_price
				#print("总价-------")
				#print(total_price)
			#得到单价，写入文件中
			elif node.get_text()=='单价：':
				unit_price = node.find_next_sibling().get_text()
				unit_price = "".join(unit_price.split())
				#dataSheet.cell(maxRow+1,5,unit_price)
				#单价写入关键信息行
				linkDetailInfoRow[3]=unit_price
				#print("单价-------")
				#print(unit_price)
				
			#得到预期租金收益，写入文件中
			elif node.get_text()=='预期：':
				rent = node.find_next_sibling().get_text()
				rent = "".join(rent.split())
				if rent!='暂无数据':
					#写入是，去除文本中的 元/月，并转换为int类型
					rent = int(rent[:-7])
					#dataSheet.cell(maxRow+1,3,rent)
					#预期租金收益写入关键信息行
					linkDetailInfoRow[4]=rent
					#print("预期-------")
					#print(rent)	
		
		#上传日期
		releaseDate = self.getReleaseDate(soup_childPage)
		#dataSheet.cell(maxRow+1,7,releaseDate)
		#上传日期写入关键信息行
		linkDetailInfoRow[5]=releaseDate
		
		#备注信息
		comment = self.getComment(soup_childPage)
		#dataSheet.cell(maxRow+1,12,comment)
		#备注信息写入关键信息行
		linkDetailInfoRow[6]=comment
		
		#面积信息
		area = self.getArea(soup_childPage)
		#dataSheet.cell(maxRow+1,6,area)
		#面积信息写入关键信息行
		linkDetailInfoRow[7]=area
		
		#wb.save('安居客爬取数据.xlsx')
		
		#返回该行关键信息存储顺序为，链接，大概地理位置，总价，单价，预期租金收益，上传日期，备注信息，面积

		#print(linkDetailInfoRow)
		return linkDetailInfoRow

		
		
	#获取页面关于该门面的面积信息				
	def getArea(self,soup_childPage):
		area = soup_childPage.find('ul','ritem').find('span','desc').get_text()
		#print("建筑面积---------")
		#print(area)
		return area
		
		
	
	#获取页面关于该门面的详细信息				
	def getComment(self,soup_childPage):
		childPageComment = soup_childPage.find('div','desc-con').get_text().strip()
		#正则表达是，根据多个空格隔断
		info = re.compile('\s+')
		#多个空格的部分转化为跳行
		splitedChildPageComment = info.sub('\n',childPageComment)
		
		#print("详细描述---------")
		#print(splitedChildPageComment)
		return splitedChildPageComment
		
		
	#获取该信息的上传日期
	def getReleaseDate(self,soup_childPage):
		#releaseDate = soup_childPage.find('div','hd-sub').get_text()[5:15]
		releaseDate = soup_childPage.find('div','hd-sub').get_text()
		releaseDate = "".join(releaseDate.split())
		#print("上传时间---------")
		#print(releaseDate[4:14])
		return releaseDate[4:14]
		
		


# 运行方法2：

if __name__ == '__main__':
	
	#fakeHeadersList = Disguise.Disguise().generateFakeHeadersList(20)
	
	#方法1：通过地点爬取数据
	def runAnjukeSpiderByLocation():
		spider = AnjukeSpider()
		#locationList = ["jiangbei","yubei","yuzhong","shangpingba","jiulongpo","nananqu","dadukou","banan","beibei"]
		locationList = ["yubei"]
		#数字代表要爬取的总页数
		localLinksSet = spider.loadLocalData()
		for location in locationList:
			for pageNum in range(1,2):
				list_items = spider.goToMainPage(pageNum,location)
				wholePageLinklinkDetailInfoRows = spider.getWholePageNewInfo(list_items,localLinksSet)
				spider.writeWholePageInfoAndSave(wholePageLinklinkDetailInfoRows)
	
	
	#方法2：通过关键词爬取数据
	def runAnjukeSpiderByKeyword():
		spider = AnjukeSpider()
		#keywordList = [”月租“,"磁器口",”西站“,”租约“,”包税“,”银行“,”南开“,”包干“,”洪崖洞“,”肯德基“,”麦当劳“,”“,”“,”“,”“,”“,”“,”“,”“,”“,”“,”“,”“,”“,”“,”“,”“,”“,]
		keywordList = ["月租"]
		localLinksSet = spider.loadLocalData()
		for word in keywordList:
			for pageNum in range(33,100):
				list_items = spider.goToMainPage(pageNum,keyword=word)
				wholePageLinklinkDetailInfoRows = spider.getWholePageNewInfo(list_items,localLinksSet)
				spider.writeWholePageInfoAndSave(wholePageLinklinkDetailInfoRows)
	
	
	#直接运行
	#runAnjukeSpiderByLocation()
	runAnjukeSpiderByKeyword()