import jieba       #中文分词
import openpyxl    #读写excel 


'''
读取爬取的详细信息，利用中文分词，提取月租金和年递增率信息
并计算周期，月数
'''

class AnalyzeRentRatio:
	

	def readAndAnalyzeFile(self):
		
		#读取爬取的数据
		wb = openpyxl.load_workbook('安居客爬取数据.xlsx')
		dataSheet = wb["重庆"]
		nrows = dataSheet.max_row
		#缓存list，用两个空的list占位，因为读写表格从第2行开始
		allRentRatioPairs = [[0,0],[0,0]]
		for i in range(2,nrows+1):
		#for i in range(2,15):
			print('--------第',i,'个')
			ss=''
			#读取详细信息
			ss1=dataSheet.cell(i,12).value
			#处理空格和换行符
			#ss2=ss1.replace(' ','').replace('\n','')
			#仅处理空格
			#print('SS1:'+ss1)
			ss2=ss1.replace(' ','')
			ss+=ss2
			#利用jieba分词，用list存储
			seg_list = jieba.lcut(ss, cut_all=False)
			#print("/ ".join(seg_list))
			rentRatioPair = self.getRentAndRatioPair(seg_list)
			#print(rentRatioPair)
			allRentRatioPairs.append(rentRatioPair)
			
		return allRentRatioPairs
			
	def getRentAndRatioPair (self,seg_list):
		
		#创建缓存list存取单条信息的租金和递增率,第一个存租金，第二个存递增率
		rentRatioPair = [0,0]
		#遍历分词list
		for word_index in range(len(seg_list)):
			#如果词为list中的关键词，下一个或下下个词为数字，则为租金数额
			if seg_list[word_index] in ['月租','月租金','租金','租']:
				if word_index+1<len(seg_list) and seg_list[word_index+1].isdigit() :
					if int(seg_list[word_index+1])>1000:
						#print('月租：-------',seg_list[word_index+1])
						rent = seg_list[word_index+1]
						rentRatioPair[0] = int(rent)
				elif word_index+2<len(seg_list) and seg_list[word_index+2].isdigit():
					if int(seg_list[word_index+2])>1000:
						#print('月租：-------',seg_list[word_index+2])
						rent = seg_list[word_index+2]
						rentRatioPair[0] = int(rent)
							
						
			#如果词为list中的关键词，下一个或下下个词为数字，则为递增率信息		
			elif seg_list[word_index] == '递增':
				ratio = '0'
				if word_index+1<len(seg_list) and ('%'in seg_list[word_index+1]):
					#print('递增率：-------',seg_list[word_index+1][:-1])
					ratio = seg_list[word_index+1][:-1]
					
				elif word_index+1<len(seg_list) and seg_list[word_index+1]=='百分之':
					#print('递增率：-------',seg_list[word_index+2])
					ratio = seg_list[word_index+2]

				#else:
					#print('递增率：-------',0)
				if ratio.isdigit():
					rentRatioPair[1] = int(ratio)
					
		return rentRatioPair
		
		
		
	def writeToFile(self):
		#将缓存写入文件中
		
		allRentRatioPairs = self.readAndAnalyzeFile()
		wb = openpyxl.load_workbook('安居客爬取数据.xlsx')
		dataSheet = wb["重庆"]
		nrows = dataSheet.max_row
		for i in range(2,nrows+1):
		#for i in range(2,15):
			if dataSheet.cell(i,3).value == None:
				dataSheet.cell(i,3,allRentRatioPairs[i][0])
			if dataSheet.cell(i,4).value == None:
				dataSheet.cell(i,4,allRentRatioPairs[i][1])
		wb.save('安居客爬取数据.xlsx')
		
		
		

		
	def calPaybackPeriod(self):
		#计算回报周期，并写入文件中
		
		
		#读取爬取的数据
		wb = openpyxl.load_workbook('安居客爬取数据.xlsx')
		dataSheet = wb["重庆"]
		nrows = dataSheet.max_row
		#缓存list，用两个空的list占位，因为读写表格从第2行开始
		totalMonthList = [0,0]
		for i in range(2,nrows+1):
			print("第",i,"个")
			
			#总价，如12万 转换为120000
			totalPrice=round(float(dataSheet.cell(i,2).value[:-1]))*10000
			print(totalPrice)
			#如果租金为0，则不继续计算，跳入下一个
			rent = dataSheet.cell(i,3).value
			if rent ==0:
				totalMonthList.append(0)
				print()
				continue
				
			#如果递增率为0，则设置为默认0.05	
			ratio = dataSheet.cell(i,4).value
			if ratio == 0:
				rate = 0.05
			else:
				rate = float(ratio/100)
			
			#计算回报周期		
			totalRent = 0
			totalMonth = 0
			totalYear = -1 
				
			for i in range(1000):
				j = i//12
				
				monthRent = rent*((1+rate)**j)
				yearRent = rent*((1+rate)**j)*12
				totalRent += monthRent
				totalMonth += 1
				
				if totalYear<j:
					totalYear = j	
				if totalRent > totalPrice:
						break

			print("共需" + str(totalMonth)+ "个月回本")
			print()
			totalMonthList.append(totalMonth)
		
		
		#写入文件中
		wb = openpyxl.load_workbook('安居客爬取数据.xlsx')
		dataSheet = wb["重庆"]
		nrows = dataSheet.max_row
		for i in range(2,nrows+1):
			dataSheet.cell(i,11,totalMonthList[i])
			
		wb.save('安居客爬取数据.xlsx')
		#return totalMonthList	



readRatio = AnalyzeRentRatio()
#result = readRatio.readAndAnalyzeFile()
#print(result)
readRatio.writeToFile()

readRatio.calPaybackPeriod()
